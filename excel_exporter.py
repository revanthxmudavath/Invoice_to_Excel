"""
Excel export module for invoice data
"""

import logging
from pathlib import Path
from typing import Dict, Any, List, Optional
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

class ExcelExporter:
    """Exports invoice data to Excel format"""
    
    def __init__(self):
        self.logger = logging.getLogger(__name__)
        
        # Define styles
        self.header_font = Font(bold=True, color="FFFFFF")
        self.header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        self.subheader_font = Font(bold=True, color="000000")
        self.subheader_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        self.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
    
    def export_invoice(self, invoice_data: Dict[str, Any], output_path: str) -> str:
        """
        Export invoice data to Excel file
        
        Args:
            invoice_data: Parsed invoice data
            output_path: Path for the Excel file
            
        Returns:
            Path to the created Excel file
        """
        try:
            self.logger.info(f"Starting Excel export for invoice: {invoice_data.get('invoice_number', 'Unknown')}")
            
            # Create workbook
            wb = openpyxl.Workbook()
            
            # Remove default sheet
            wb.remove(wb.active)
            
            # Create sheets
            self.logger.info("Creating summary sheet...")
            self._create_summary_sheet(wb, invoice_data)
            
            self.logger.info("Creating items sheet...")
            self._create_items_sheet(wb, invoice_data)
            
            self.logger.info("Creating vendor sheet...")
            self._create_vendor_sheet(wb, invoice_data)
            
            self.logger.info("Creating validation sheet...")
            self._create_validation_sheet(wb, invoice_data)
            
            # Save workbook
            wb.save(output_path)
            wb.close()
            
            self.logger.info(f"Excel file exported successfully to: {output_path}")
            return output_path
            
        except Exception as e:
            self.logger.error(f"Error exporting to Excel: {e}")
            import traceback
            self.logger.error(f"Traceback: {traceback.format_exc()}")
            raise
    
    def _create_summary_sheet(self, wb: openpyxl.Workbook, data: Dict[str, Any]):
        """Create summary sheet with key invoice information"""
        ws = wb.create_sheet("Invoice Summary")
        
        # Set column widths
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 40
        
        # Title
        ws['A1'] = "INVOICE SUMMARY"
        ws['A1'].font = Font(size=16, bold=True)
        ws.merge_cells('A1:B1')
        
        # Invoice details
        row = 3
        summary_fields = [
            ("Invoice Number", data.get("invoice_number")),
            ("Invoice Date", data.get("invoice_date")),
            ("Vendor", data.get("vendor_name")),
            ("Customer Number", data.get("customer_number")),
            ("Route", data.get("route")),
            ("Stop", data.get("stop")),
            ("Terms", data.get("terms")),
            ("Due Date", data.get("due_date")),
            ("PO Number", data.get("po_number")),
            ("License", data.get("license")),
            ("Exp Date", data.get("exp_date")),
            ("Chain", data.get("chain")),
            ("Delivery Number", data.get("delivery_number")),
        ]
        
        for field, value in summary_fields:
            if value is not None and value != "" and value != "None":
                ws[f'A{row}'] = field
                ws[f'B{row}'] = str(value)
                ws[f'A{row}'].font = self.subheader_font
                ws[f'A{row}'].fill = self.subheader_fill
                ws[f'A{row}'].border = self.border
                ws[f'B{row}'].border = self.border
                row += 1
        
        # Totals section
        row += 1
        ws[f'A{row}'] = "TOTALS"
        ws[f'A{row}'].font = Font(size=14, bold=True)
        ws[f'A{row}'].fill = PatternFill(start_color="FFD966", end_color="FFD966", fill_type="solid")
        ws.merge_cells(f'A{row}:B{row}')
        row += 1
        
        total_fields = [
            ("Total Bottles", data.get("total_bottles")),
            ("Total Liquor Gallons", data.get("total_liquor_gallons")),
            ("Total Beer Gallons", data.get("total_beer_gallons")),
            ("Gross Total", data.get("gross_total")),
            ("Total Discount", data.get("total_discount")),
            ("Net Amount", data.get("net_amount")),
        ]
        
        for field, value in total_fields:
            if value is not None and value != "" and value != "None":
                ws[f'A{row}'] = field
                ws[f'B{row}'] = value
                ws[f'A{row}'].font = self.subheader_font
                ws[f'A{row}'].fill = self.subheader_fill
                ws[f'A{row}'].border = self.border
                ws[f'B{row}'].border = self.border
                row += 1
    
    def _create_items_sheet(self, wb: openpyxl.Workbook, data: Dict[str, Any]):
        """Create items sheet with detailed line items"""
        ws = wb.create_sheet("Items")
        
        # Set column widths
        column_widths = [8, 12, 35, 15, 15, 15, 15, 15, 15, 15, 15, 15, 15, 15]
        for i, width in enumerate(column_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = width
        
        # Headers
        headers = [
            "Case", "Btles", "Item", "Size", "BPC", "Description", 
            "cs_price", "cs_disc", "cs_net", "cnty_tax", "city_tax", 
            "ext_w/o_tax", "slp", "deal"
        ]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.border = self.border
            cell.alignment = Alignment(horizontal="center")
        
        # Data rows
        items = data.get("items", [])
        for row_idx, item in enumerate(items, 2):
            for col_idx, header in enumerate(headers, 1):
                value = item.get(header, "")
                # Handle None and "None" values
                if value is None or value == "None":
                    value = ""
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.border = self.border
                
                # Format numeric columns
                if header in ["cs_price", "cs_disc", "cs_net", "cnty_tax", "city_tax", "ext_w/o_tax"]:
                    if value and str(value).replace('.', '').replace('-', '').isdigit():
                        cell.number_format = '#,##0.00'
        
        # Auto-filter
        ws.auto_filter.ref = ws.dimensions
    
    def _create_vendor_sheet(self, wb: openpyxl.Workbook, data: Dict[str, Any]):
        """Create vendor information sheet"""
        ws = wb.create_sheet("Vendor Info")
        
        # Set column widths
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 50
        
        # Title
        ws['A1'] = "VENDOR INFORMATION"
        ws['A1'].font = Font(size=16, bold=True)
        ws.merge_cells('A1:B1')
        
        # Vendor details
        row = 3
        vendor_fields = [
            ("Vendor Name", data.get("vendor_name")),
            ("Vendor Address", data.get("vendor_address")),
            ("Vendor Phone", data.get("vendor_phone")),
            ("Remit To Address", data.get("remit_to_address")),
            ("Barcode", data.get("barcode")),
            ("Special Instructions", data.get("special_instructions")),
        ]
        
        for field, value in vendor_fields:
            if value is not None and value != "" and value != "None":
                ws[f'A{row}'] = field
                ws[f'B{row}'] = str(value)
                ws[f'A{row}'].font = self.subheader_font
                ws[f'A{row}'].fill = self.subheader_fill
                ws[f'A{row}'].border = self.border
                ws[f'B{row}'].border = self.border
                row += 1
    
    def _create_validation_sheet(self, wb: openpyxl.Workbook, data: Dict[str, Any]):
        """Create validation and metadata sheet"""
        ws = wb.create_sheet("Validation")
        
        # Set column widths
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 50
        
        # Title
        ws['A1'] = "VALIDATION & METADATA"
        ws['A1'].font = Font(size=16, bold=True)
        ws.merge_cells('A1:B1')
        
        # Metadata
        row = 3
        meta = data.get("meta", {})
        if meta is None:
            meta = {}
        meta_fields = [
            ("Source File", meta.get("source_file", "")),
            ("Vendor Detected", meta.get("vendor_detected", "")),
            ("Parse Confidence", meta.get("parse_confidence", "")),
            ("Validation Flags", ", ".join(meta.get("validation_flags", []))),
            ("Export Date", datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
        ]
        
        for field, value in meta_fields:
            if value is not None and value != "" and value != "None":
                ws[f'A{row}'] = field
                ws[f'B{row}'] = str(value)
                ws[f'A{row}'].font = self.subheader_font
                ws[f'A{row}'].fill = self.subheader_fill
                ws[f'A{row}'].border = self.border
                ws[f'B{row}'].border = self.border
                row += 1
        
        # Validation summary
        row += 1
        ws[f'A{row}'] = "VALIDATION SUMMARY"
        ws[f'A{row}'].font = Font(size=14, bold=True)
        ws[f'A{row}'].fill = PatternFill(start_color="FFD966", end_color="FFD966", fill_type="solid")
        ws.merge_cells(f'A{row}:B{row}')
        row += 1
        
        # Count items
        items_count = len(data.get("items", []))
        ws[f'A{row}'] = "Total Items Extracted"
        ws[f'B{row}'] = items_count
        ws[f'A{row}'].font = self.subheader_font
        ws[f'A{row}'].fill = self.subheader_fill
        ws[f'A{row}'].border = self.border
        ws[f'B{row}'].border = self.border
    
    def export_batch(self, invoices: List[Dict[str, Any]], output_dir: str, vendor: str) -> str:
        """
        Export multiple invoices to a single Excel file with multiple sheets
        
        Args:
            invoices: List of parsed invoice data
            output_dir: Output directory
            vendor: Vendor type
            
        Returns:
            Path to the created Excel file
        """
        try:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            output_path = Path(output_dir) / f"batch_{vendor}_{timestamp}.xlsx"
            
            # Create workbook
            wb = openpyxl.Workbook()
            
            # Remove default sheet
            wb.remove(wb.active)
            
            # Create summary sheet for all invoices
            self._create_batch_summary_sheet(wb, invoices, vendor)
            
            # Create individual invoice sheets
            for i, invoice in enumerate(invoices):
                sheet_name = f"Invoice_{i+1}_{invoice.get('invoice_number', 'Unknown')}"
                sheet_name = sheet_name[:31]  # Excel sheet name limit
                
                # Create a copy of the invoice data for this sheet
                invoice_copy = invoice.copy()
                invoice_copy["meta"] = invoice_copy.get("meta", {})
                invoice_copy["meta"]["batch_position"] = i + 1
                
                # Create the sheet
                self._create_summary_sheet(wb, invoice_copy)
                wb.active.title = sheet_name
            
            # Save workbook
            wb.save(output_path)
            wb.close()
            
            self.logger.info(f"Batch Excel file exported successfully to: {output_path}")
            return str(output_path)
            
        except Exception as e:
            self.logger.error(f"Error exporting batch to Excel: {e}")
            raise
    
    def _create_batch_summary_sheet(self, wb: openpyxl.Workbook, invoices: List[Dict[str, Any]], vendor: str):
        """Create summary sheet for batch processing"""
        ws = wb.create_sheet("Batch Summary")
        
        # Set column widths
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 20
        
        # Title
        ws['A1'] = f"BATCH PROCESSING SUMMARY - {vendor.upper()}"
        ws['A1'].font = Font(size=16, bold=True)
        ws.merge_cells('A1:E1')
        
        # Headers
        headers = ["Invoice #", "Date", "Items", "Total Amount", "Status"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.border = self.border
            cell.alignment = Alignment(horizontal="center")
        
        # Data rows
        for row_idx, invoice in enumerate(invoices, 4):
            ws.cell(row=row_idx, column=1, value=invoice.get("invoice_number", "N/A"))
            ws.cell(row=row_idx, column=2, value=invoice.get("invoice_date", "N/A"))
            ws.cell(row=row_idx, column=3, value=len(invoice.get("items", [])))
            
            total = invoice.get("net_amount") or invoice.get("gross_total") or invoice.get("total_sales")
            ws.cell(row=row_idx, column=4, value=total or "N/A")
            if total and str(total).replace('.', '').replace('-', '').isdigit():
                ws.cell(row=row_idx, column=4).number_format = '#,##0.00'
            
            # Status based on validation flags
            flags = invoice.get("meta", {}).get("validation_flags", [])
            if not flags:
                status = "✅ Valid"
            else:
                status = f"⚠️ {len(flags)} flags"
            
            ws.cell(row=row_idx, column=5, value=status)
            
            # Apply borders
            for col in range(1, 6):
                ws.cell(row=row_idx, column=col).border = self.border
        
        # Auto-filter
        ws.auto_filter.ref = ws.dimensions
        
        # Summary stats
        row = len(invoices) + 5
        ws[f'A{row}'] = "BATCH STATISTICS"
        ws[f'A{row}'].font = Font(size=14, bold=True)
        ws[f'A{row}'].fill = PatternFill(start_color="FFD966", end_color="FFD966", fill_type="solid")
        ws.merge_cells(f'A{row}:E{row}')
        
        row += 1
        ws[f'A{row}'] = f"Total Invoices: {len(invoices)}"
        ws[f'B{row}'] = f"Total Items: {sum(len(inv.get('items', [])) for inv in invoices)}"
        
        # Calculate total amount
        total_amount = 0
        for invoice in invoices:
            amount = invoice.get("net_amount") or invoice.get("gross_total") or invoice.get("total_sales")
            if amount and str(amount).replace('.', '').replace('-', '').isdigit():
                total_amount += float(amount)
        
        ws[f'C{row}'] = f"Total Amount: ${total_amount:,.2f}"
